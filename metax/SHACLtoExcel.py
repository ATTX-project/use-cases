from openpyxl import Workbook
from rdflib import Namespace, Graph, URIRef
from rdflib.namespace import RDF, RDFS, DCTERMS


def getLabel(subject, lang):
    labels = g.objects(subject,RDFS.label)
    for l in labels:
        if l.language == lang:
            return unicode(l)



def getValues(subject, predicate):
    value = ''
    values = data.objects(subject,predicate)
    for v in values:
        value = unicode(v) + '|'

    return value
# Load ATT SHACL description
g=Graph()
data=Graph()
g.parse("att-shacl.ttl", format="turtle")
data.parse("input-data.ttl", format="turtle")

ns_sh = Namespace("http://www.w3.org/ns/shacl#")
ns_att = Namespace("http://iow.csc.fi/ns/att#")
ns_owl = Namespace("http://www.w3.org/2002/07/owl#")

wb = Workbook()

sheets = []

for s,p,o in g.triples( (None, RDF.type, ns_sh['Shape'] ) ):
    shapeLabel = getLabel(s, 'fi')
    sheets.append({'label': shapeLabel, 'subject': s})

sheets = sorted(sheets, key=lambda k: k['label'])

for sheet in sheets:
    ws = wb.create_sheet(sheet['label'])
    s = sheet['subject']
    scopeClass = g.value(s, ns_sh['scopeClass'])
    if scopeClass == None:
        scopeClass = g.value(s, ns_sh['targetClass'])



    columnData = []
    ws.cell(row=1, column=1, value="URI")
    for s2, p2, o2 in g.triples( (s, ns_sh['property'], None) ):
        propertyLabel = getLabel(o2, 'fi')
        predicate = g.value(o2, ns_sh['predicate'])
        if predicate == None:
            predicate = g.value(o2, ns_sh['path'])

        predType = g.value(o2, DCTERMS.type)
        minCount = g.value(o2, ns_sh['minCount'])
        index = g.value(o2, ns_sh['index'])
        if minCount:
            minCount = int(minCount)

        if index:
            index = int(index)
        else:
            index = 0

        if unicode(predType) == 'http://www.w3.org/2002/07/owl#ObjectProperty':
            propertyLabel = propertyLabel + " (URI)"
        if(minCount > 0):
            propertyLabel = propertyLabel + " *"

        columnData.append({'index': index, 'predicate': URIRef(predicate), 'label': propertyLabel})

    columnData = sorted(columnData, key=lambda k: k['index'])
    col = 2
    for c in columnData:
        ws.cell(row=1, column=col, value=c['label'])
        col = col + 1

    # loop through values
    row = 2
    subjects = data.subjects(predicate=RDF.type, object=URIRef(scopeClass))
    for subject in subjects:
        col = 2
        ws.cell(row=row, column=1, value=unicode(subject))
        for cd in columnData:
            pred = cd['predicate']
            values = getValues(subject, pred)
            ws.cell(row=row, column=col, value=values)

            col = col + 1
        row = row + 1


wb.save('output.xlsx')
