from openpyxl import Workbook
from rdflib import Namespace, Graph, URIRef
from rdflib.namespace import RDF, RDFS, DCTERMS


def getLabel(subject, lang):
    labels = g.objects(subject,RDFS.label)
    for l in labels:
        if l.language == lang:
            return unicode(l)

def getFirstValue(subject, predicate):
    value = ''
    values = data.objects(subject,predicate)
    for v in values:
        return unicode(v)

    return ''

def getValues(subject, predicate):
    value = ''
    values = data.objects(subject,predicate)
    for v in values:
        value = unicode(v) + '|'

    return value
# Load ATT SHACL description
data=Graph()
data.parse("zenodo-uh-base-output.ttl", format="turtle")
#data.parse("etsin-uh-output.ttl", format="turtle")
#data.parse("yso-skos.ttl", format="turtle")
#data.parse("ysa-skos.ttl", format="turtle")

ns_sh = Namespace("http://www.w3.org/ns/shacl#")
ns_att = Namespace("http://iow.csc.fi/ns/att#")
ns_owl = Namespace("http://www.w3.org/2002/07/owl#")
ns_attx = Namespace("http://data.hulib.helsinki.fi/attx/work#")
ns_skos = Namespace("http://www.w3.org/2004/02/skos/core#")

wb = Workbook()

ws = wb.active
ws.title = "Data"

ws['A1'] = "identifier"
ws['B1'] = "creators"
ws['C1'] = "contributors"
ws['D1'] = "title"
ws['E1'] = "publisher"
ws['F1'] = "publicationYear"
ws['G1'] = "resourceType"
ws['H1'] = "subjects"
ws['I1'] = "language"
ws['J1'] = "rights"
ws['K1'] = "description"


i = 2
for s,p,o in data.triples( (None, RDF.type, ns_attx['Dataset'] ) ):
    # identifier
    # - identifier
    # - identifierType
    content = ''
    for identifierResource in data.objects(s, ns_attx['identifier']):
        identifier = getFirstValue(identifierResource, ns_attx['value'])
        identifierType = getFirstValue(identifierResource, ns_attx['type'])

        content = identifier + ' (' + identifierType + ')'

    print content
    ws['A' + str(i)] = content

    # creators
    # - creatorName
    # - affiliations
    # - nameIdentifiers
    #   - nameIdentifier
    #   - nameIdentifierScheme
    content = ''
    for r in data.objects(s, ns_attx['creator']):
        name = getFirstValue(r, ns_attx['creatorName'])
        aff = getFirstValue(r, ns_attx['affiliation'])
        identifier = getFirstValue(r, ns_attx['nameIdentifier'])

        content = content  + name + ', ' + aff + ' (' + identifier + ')|'

    print content
    ws['B' + str(i)] = content

    # contributors
    # - contributorName
    # - type
    # - affiliations
    # - nameIdentifiers
    #   - nameIdentifier
    #   - nameIdentifierScheme
    content = ''
    for r in data.objects(s, ns_attx['contributor']):
        name = getFirstValue(r, ns_attx['contributorName'])
        typeName = getFirstValue(r, ns_attx['contributorType'])
        aff = getFirstValue(r, ns_attx['affiliation'])
        identifier = getFirstValue(r, ns_attx['nameIdentifier'])

        content = content  + typeName + ': ' + name + ' ' + aff + ' (' + identifier + ')|'

    print content
    ws['C' + str(i)] = content


    # title
    content = getValues(s, ns_attx['title'])
    print content
    ws['D' + str(i)] = content


    # publisher
    content = getValues(s, ns_attx['publisher'])
    print content
    ws['E' + str(i)] = content

    # publicationYear
    content = getValues(s, ns_attx['pubdate'])
    print content
    ws['F' + str(i)] = content

    # resource type
    content = ''
    for r in data.objects(s, ns_attx['resourceType']):
        t = getFirstValue(r, ns_attx['type'])
        title = getFirstValue(r, ns_attx['title'])
        content = content  + title + ' (' + t + ')'

    print content
    ws['G' + str(i)] = content

    # subjects
    # - subject
    # - scheme
    # - valueURI
    content = ''
    for r in data.objects(s, ns_attx['subject']):
        label = getFirstValue(r, ns_attx['term'])
        if label == '':
            label = getFirstValue(r, ns_skos['prefLabel'])

        scheme = getFirstValue(r, ns_attx['subjectScheme'])
        if scheme == '':
            scheme = getFirstValue(r, ns_skos['inScheme'])

        identifier = getFirstValue(r, ns_attx['valueURI'])
        if identifier == '':
            identifier = str(r)
        content = content  + label + ' (' + scheme + ':' + identifier + ')|'
    print content
    ws['H' + str(i)] = content


    # language
    content = getValues(s, ns_attx['pubdate'])
    print content
    ws['I' + str(i)] = content

    # rightsList
    # - rights
    # - rights uri
    content = ''
    for r in data.objects(s, ns_attx['license']):
        label = getFirstValue(r, ns_attx['licenseID'])
        identifier = getFirstValue(r, ns_attx['valueURI'])
        content = content  + label + ' (' + identifier + ')|'
    print content
    ws['J' + str(i)] = content

    # description
    content = getValues(s, ns_attx['description'])
    print content
    ws['K' + str(i)] = content
    i = i + 1

#wb.save('etsin-output.xlsx')
wb.save('zenodo-output.xlsx')
